import openpyxl
import pandas as pd
import os

#loading excel sheet >> DONE
def load_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    else:
        return "File Not Found :("

wb_path = "data.xlsx"
wb = load_workbook(wb_path)
sheet = wb["Sheet"]
sheet_obj = wb.active
max_column = sheet_obj.max_column
max_row = sheet_obj.max_row

#main menu >> DONE
def menu():
    
    #create item >> DONE
    def create_item():
        id = int(input("\nEnter ID\n"))
        name = input("\nEnter Name\n")
        amount = int(input("\nEnter Amount\n"))
        sheet.append([id, name, amount])
        wb.save(wb_path)
        print("Item Added Succesfully")
        add_more = input("\nWants To Add More Item? Y/N\n")
        if add_more.lower() == "yes":
            create_item()

    #view item >> DONE
    def view_item():
        item_list = pd.read_excel(wb_path)
        print(item_list)

    #search item
    def search_item(id):
        for i in range(1,max_row+1):
            if sheet.cell(row=i,column = 1).value == id:
                print("Item Found")
                return i

    #display item
    def display_item(row):
        for i in range(1,max_column+1):
            cell_obj = sheet_obj.cell(row = row, column = i)
            print(cell_obj.value)

    #Update item
    def update_item(row):
        x = int(input('\n Input ID of Item :'))
        for col_index,value in enumerate(x,start =1):
            sheet.cell(row= row, column=col_index,value=value)
        print("\n Item has been Updated") 

    #delete item
    def delete_item(row):
        sheet.delete_rows(row)
        wb.save(wb_path)
        print("Item Has Been Deleted")        
    
    while True:
        print("Warehouse Administration Program")
        print("\nProgram By :")
        print("ZakariyaXD x Midnight SpeedBoyz")
        print("\nItem Data List\n")
        print("1. Create New Item Data")
        print("2. View Item Data")
        print("3. Update Item Data")
        print("4. Delete Item")
        print("0. Exit The Program")
        ch = input("\nEnter the option\n")
        if ch == '1':
            create_item()
        elif ch == '2':
            view_item()
        elif ch == '3':
            x = int(input("Enter ID"))
            row = search_item(x)
            display_item(row)
            y = input("Edit? Y/N")
            if y == "y":
                update_item(row)
        elif ch == '4':
            x = int(input("Enter ID"))
            row = search_item(x)
            display_item(row)
            y = input("Delete? Y/N")
            if y == 'yes':
                delete_item()

        elif ch == '0':
            exit_program = input("Are You Sure? Y/N\n")
            if exit_program.lower() == "y":
                print(exit)
                os._exit(0)
            else:
                menu()
        else:
            print("Wrong Command")



#            def by_id(id):
#               for i in range(1,max_row):
#                    if sheet.cell(row=i,column = 1).value == id:
#                        print("Item Found")
#                        return i
#                     else:
#                        print("Item Not Found :(\n")
#                        search_item()
#                        
#            def by_name(name):
#                for n in range(1,max_row):
#                    if sheet.cell(row=n,column = 2).value == name:
#                        print("Item Found")
#                        return n
#                    else:
#                        print("Item Not Found")
#                        search_item()
while True:
    menu()